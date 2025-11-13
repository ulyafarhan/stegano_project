from .interface import SteganographyStrategy
import io
from openpyxl import load_workbook
import base64

class XlsxHandler(SteganographyStrategy):
    METADATA_FIELD = "description" # Menggunakan 'description'
    MAX_CAPACITY_BYTES = 60 * 1024 # Batas aman 60KB

    def check_capacity(self, host_file: io.BytesIO, payload_size: int) -> bool:
        return payload_size <= self.MAX_CAPACITY_BYTES

    def encode(self, host_file: io.BytesIO, payload: bytes) -> io.BytesIO:
        host_file.seek(0)
        encoded_payload = base64.b64encode(payload).decode('utf-8')
        
        try:
            wb = load_workbook(host_file)
        except Exception:
             raise ValueError("File .xlsx tidak valid atau rusak.")
        
        props = wb.properties
        props.description = encoded_payload
        
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer

    def decode(self, host_file: io.BytesIO) -> bytes:
        host_file.seek(0)
        try:
            wb = load_workbook(host_file)
        except Exception:
             raise ValueError("File .xlsx tidak valid atau rusak.")
        
        props = wb.properties
        encoded_payload = props.description
        
        if not encoded_payload:
            raise ValueError("Tidak ada data steganografi (metadata) ditemukan.")
        
        try:
            return base64.b64decode(encoded_payload)
        except Exception:
            raise ValueError("Data metadata rusak atau format base64 salah.")